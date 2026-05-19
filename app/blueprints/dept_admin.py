from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from werkzeug.security import generate_password_hash
from functools import wraps
from app.db import get_db
from app.utils import log_action, generate_unit_report_csv
import secrets
import string

dept_admin_bp = Blueprint('dept_admin', __name__)


def dept_admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ('dept_admin', 'super_admin'):
            flash('Department Admin access required.', 'danger')
            return redirect(url_for('auth.login'))
        return f(*args, **kwargs)
    return decorated


def generate_temp_password(length=10):
    chars = string.ascii_letters + string.digits + '!@#$'
    return ''.join(secrets.choice(chars) for _ in range(length))


# ─────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/dashboard')
@login_required
@dept_admin_required
def dashboard():
    db     = get_db()
    dep_id = str(current_user.department_id) if current_user.department_id else None
    stats  = {}
    try:
        q_cls = db.table('classes').select('id')
        if dep_id:
            q_cls = q_cls.eq('department_id', dep_id)
        stats['classes'] = len(q_cls.execute().data or [])

        q_trainers = db.table('users').select('id').eq('role', 'trainer')
        q_trainees = db.table('users').select('id').eq('role', 'trainee')
        if dep_id:
            q_trainers = q_trainers.eq('department_id', dep_id)
            q_trainees = q_trainees.eq('department_id', dep_id)
        stats['trainers'] = len(q_trainers.execute().data or [])
        stats['trainees'] = len(q_trainees.execute().data or [])

        # Count units via courses in this department
        course_ids = []
        if dep_id:
            course_ids = [c['id'] for c in
                          db.table('courses').select('id').eq('department_id', dep_id).execute().data or []]
        if course_ids:
            stats['units'] = len(
                db.table('units').select('id').in_('course_id', course_ids).execute().data or [])
        else:
            stats['units'] = len(db.table('units').select('id').execute().data or [])

        recent = (db.table('assessments')
                  .select('*, users!assessments_trainee_id_fkey(full_name, admission_no), units(name), classes(name)')
                  .order('uploaded_at', desc=True)
                  .limit(10)
                  .execute().data or [])
        for a in recent:
            ev = db.table('evidence').select('id').eq('assessment_id', a['id']).execute().data or []
            a['evidence_count'] = len(ev)
            from app.utils import get_storage_public_url, STORAGE_BUCKET_SCRIPTS
            a['script_url'] = get_storage_public_url(STORAGE_BUCKET_SCRIPTS, a.get('script_file_path', ''))

        # Units list for CSV download
        course_ids = []
        if dep_id:
            course_ids = [c['id'] for c in
                          db.table('courses').select('id').eq('department_id', dep_id).execute().data or []]

        if course_ids:
            units_list = (db.table('units').select('id, name')
                          .in_('course_id', course_ids).order('name').execute().data or [])
        else:
            units_list = db.table('units').select('id, name').order('name').execute().data or []

    except Exception as e:
        flash(f'Error: {e}', 'danger')
        recent = []
        units_list = []

    return render_template('dept_admin/dashboard.html',
                           stats=stats, recent=recent, units_list=units_list)


@dept_admin_bp.route('/download-unit-report/<unit_id>')
@login_required
@dept_admin_required
def download_unit_report(unit_id):
    from flask import make_response
    db = get_db()
    try:
        dep_id   = str(current_user.department_id) if current_user.department_id else None
        csv_data = generate_unit_report_csv(unit_id, department_id=dep_id)
        unit     = db.table('units').select('name').eq('id', unit_id).single().execute().data
        name     = unit['name'].replace(' ', '_') if unit else 'Unit'
        resp     = make_response(csv_data)
        resp.headers['Content-Disposition'] = f'attachment; filename="report_{name}.csv"'
        resp.headers['Content-type'] = 'text/csv'
        return resp
    except Exception as e:
        flash(f'Error: {e}', 'danger')
        return redirect(url_for('dept_admin.dashboard'))


# ─────────────────────────────────────────────────────────────
# COURSES
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/courses')
@login_required
@dept_admin_required
def courses():
    db     = get_db()
    dep_id = str(current_user.department_id) if current_user.department_id else None
    q      = db.table('courses').select('*, departments(name)')
    if dep_id:
        q = q.eq('department_id', dep_id)
    data = q.order('name').execute().data or []
    deps = db.table('departments').select('*').order('name').execute().data or []
    return render_template('dept_admin/courses.html', courses=data, departments=deps)


@dept_admin_bp.route('/courses/add', methods=['POST'])
@login_required
@dept_admin_required
def add_course():
    db     = get_db()
    dep_id = request.form.get('department_id') or str(current_user.department_id)
    name   = request.form.get('name', '').strip()
    code   = request.form.get('code', '').strip().upper()
    if not name or not code or not dep_id:
        flash('All fields required.', 'danger')
        return redirect(url_for('dept_admin.courses'))
    try:
        db.table('courses').insert({
            'name': name, 'code': code,
            'department_id': dep_id,
            'created_by': str(current_user.id)
        }).execute()
        log_action('CREATE_COURSE', 'course', None, name)
        flash(f'Course "{name}" created.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('dept_admin.courses'))


@dept_admin_bp.route('/courses/delete/<course_id>', methods=['POST'])
@login_required
@dept_admin_required
def delete_course(course_id):
    db = get_db()
    try:
        course = db.table('courses').select('name').eq('id', course_id).single().execute().data
        name   = course['name'] if course else course_id
        db.table('courses').delete().eq('id', course_id).execute()
        log_action('DELETE_COURSE', 'course', course_id, name)
        flash(f'Course "{name}" deleted.', 'success')
    except Exception as e:
        flash(f'Error deleting course: {e}', 'danger')
    return redirect(url_for('dept_admin.courses'))


# ─────────────────────────────────────────────────────────────
# UNITS
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/units')
@login_required
@dept_admin_required
def units():
    db        = get_db()
    course_id = request.args.get('course_id', '')
    q         = db.table('units').select('*, courses(name, department_id, departments(name))')
    if course_id:
        q = q.eq('course_id', course_id)
    data    = q.order('name').execute().data or []
    courses = db.table('courses').select('*').order('name').execute().data or []
    return render_template('dept_admin/units.html', units=data, courses=courses, course_filter=course_id)


@dept_admin_bp.route('/units/add', methods=['POST'])
@login_required
@dept_admin_required
def add_unit():
    db        = get_db()
    course_id = request.form.get('course_id', '').strip()
    name      = request.form.get('name', '').strip()
    code      = request.form.get('code', '').strip().upper()
    if not course_id or not name:
        flash('Course and unit name are required.', 'danger')
        return redirect(url_for('dept_admin.units'))
    try:
        db.table('units').insert({
            'name': name, 'code': code or None,
            'course_id': course_id,
            'created_by': str(current_user.id)
        }).execute()
        log_action('CREATE_UNIT', 'unit', None, name)
        flash(f'Unit "{name}" created.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('dept_admin.units'))


@dept_admin_bp.route('/units/delete/<unit_id>', methods=['POST'])
@login_required
@dept_admin_required
def delete_unit(unit_id):
    db = get_db()
    try:
        unit = db.table('units').select('name').eq('id', unit_id).single().execute().data
        name = unit['name'] if unit else unit_id
        db.table('units').delete().eq('id', unit_id).execute()
        log_action('DELETE_UNIT', 'unit', unit_id, name)
        flash(f'Unit "{name}" deleted.', 'success')
    except Exception as e:
        flash(f'Error deleting unit: {e}', 'danger')
    return redirect(url_for('dept_admin.units'))


# ─────────────────────────────────────────────────────────────
# CLASSES
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/classes')
@login_required
@dept_admin_required
def classes():
    db     = get_db()
    dep_id = str(current_user.department_id) if current_user.department_id else None
    q      = db.table('classes').select('*, courses(name), departments(name)')
    if dep_id:
        q = q.eq('department_id', dep_id)
    data    = q.order('name').execute().data or []
    courses = db.table('courses').select('*').order('name').execute().data or []
    deps    = db.table('departments').select('*').order('name').execute().data or []
    return render_template('dept_admin/classes.html', classes=data, courses=courses, departments=deps)


@dept_admin_bp.route('/classes/add', methods=['POST'])
@login_required
@dept_admin_required
def add_class():
    db     = get_db()
    dep_id = request.form.get('department_id') or str(current_user.department_id)
    name   = request.form.get('name', '').strip()
    course_id = request.form.get('course_id', '').strip()
    if not name or not course_id or not dep_id:
        flash('All fields required.', 'danger')
        return redirect(url_for('dept_admin.classes'))
    try:
        db.table('classes').insert({
            'name':         name,
            'course_id':    course_id,
            'department_id': dep_id,
            'intake_year':  request.form.get('intake_year') or None,
            'intake_month': request.form.get('intake_month', '').strip() or None,
            'level':        request.form.get('level', '').strip() or None,
            'cycle':        request.form.get('cycle', '').strip() or None,
            'created_by':   str(current_user.id)
        }).execute()
        log_action('CREATE_CLASS', 'class', None, name)
        flash(f'Class "{name}" created.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('dept_admin.classes'))


@dept_admin_bp.route('/classes/<class_id>/units', methods=['GET', 'POST'])
@login_required
@dept_admin_required
def class_units(class_id):
    db  = get_db()
    cls = db.table('classes').select('*, courses(name)').eq('id', class_id).single().execute().data
    if not cls:
        flash('Class not found.', 'danger')
        return redirect(url_for('dept_admin.classes'))

    if request.method == 'POST':
        unit_ids = request.form.getlist('unit_ids')
        db.table('class_units').delete().eq('class_id', class_id).execute()
        for uid in unit_ids:
            try:
                db.table('class_units').insert({'class_id': class_id, 'unit_id': uid}).execute()
            except Exception:
                pass
        log_action('UPDATE_CLASS_UNITS', 'class', class_id, cls['name'])
        flash('Class units updated.', 'success')
        return redirect(url_for('dept_admin.classes'))

    all_units    = db.table('units').select('*').order('name').execute().data or []
    assigned_ids = [cu['unit_id'] for cu in
                    db.table('class_units').select('unit_id').eq('class_id', class_id).execute().data or []]
    return render_template('dept_admin/class_units.html',
                           cls=cls, all_units=all_units, assigned_ids=assigned_ids)


# ─────────────────────────────────────────────────────────────
# TRAINER MANAGEMENT
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/trainers')
@login_required
@dept_admin_required
def trainers():
    db     = get_db()
    dep_id = str(current_user.department_id) if current_user.department_id else None
    q      = db.table('users').select('*, departments(name)').eq('role', 'trainer')
    if dep_id:
        q = q.eq('department_id', dep_id)
    data  = q.order('full_name').execute().data or []
    deps  = db.table('departments').select('*').order('name').execute().data or []
    units = db.table('units').select('id, name').order('name').execute().data or []

    # Attach assigned unit ids to each trainer
    for t in data:
        assigned = (db.table('trainer_units').select('unit_id')
                    .eq('trainer_id', t['id']).execute().data or [])
        t['assigned_unit_ids'] = [a['unit_id'] for a in assigned]

    return render_template('dept_admin/trainers.html', trainers=data, departments=deps, units=units)


@dept_admin_bp.route('/trainers/add', methods=['POST'])
@login_required
@dept_admin_required
def add_trainer():
    db      = get_db()
    dep_id  = request.form.get('department_id') or str(current_user.department_id)
    email   = request.form.get('email', '').strip().lower()
    name    = request.form.get('full_name', '').strip()
    staff   = request.form.get('staff_no', '').strip()
    temp_pw = generate_temp_password()

    if not email or not name:
        flash('Email and name are required.', 'danger')
        return redirect(url_for('dept_admin.trainers'))
    try:
        db.table('users').insert({
            'email':         email,
            'full_name':     name,
            'role':          'trainer',
            'department_id': dep_id,
            'staff_no':      staff or None,
            'password_hash': generate_password_hash(temp_pw),
            'created_by':    str(current_user.id),
            'is_active':     True,
        }).execute()
        log_action('CREATE_TRAINER', 'user', None, name)
        flash(f'Trainer "{name}" created. Login: {email} | Temp password: {temp_pw}', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('dept_admin.trainers'))


@dept_admin_bp.route('/trainers/<trainer_id>/assign-units', methods=['POST'])
@login_required
@dept_admin_required
def assign_trainer_units(trainer_id):
    """Assign which units a trainer is responsible for."""
    db       = get_db()
    unit_ids = request.form.getlist('unit_ids')
    try:
        db.table('trainer_units').delete().eq('trainer_id', trainer_id).execute()
        for uid in unit_ids:
            db.table('trainer_units').insert({
                'trainer_id': trainer_id,
                'unit_id':    uid
            }).execute()
        trainer = db.table('users').select('full_name').eq('id', trainer_id).single().execute().data
        name    = trainer['full_name'] if trainer else trainer_id
        log_action('ASSIGN_TRAINER_UNITS', 'trainer', trainer_id,
                   f'{name} assigned {len(unit_ids)} unit(s)')
        flash(f'Units assigned to {name}.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('dept_admin.trainers'))


# ─────────────────────────────────────────────────────────────
# TRAINEE MANAGEMENT
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/trainees')
@login_required
@dept_admin_required
def trainees():
    db     = get_db()
    dep_id = str(current_user.department_id) if current_user.department_id else None
    q      = db.table('users').select('*, departments(name)').eq('role', 'trainee')
    if dep_id:
        q = q.eq('department_id', dep_id)
    data    = q.order('full_name').execute().data or []
    deps    = db.table('departments').select('*').order('name').execute().data or []
    classes = db.table('classes').select('*').order('name').execute().data or []

    # Enrich each trainee with their enrolled class and units
    for t in data:
        enrollments = (db.table('enrollments')
                       .select('classes(id, name)')
                       .eq('trainee_id', t['id'])
                       .execute().data or [])
        t['enrolled_classes'] = [e['classes'] for e in enrollments if e.get('classes')]

        # Collect units for each enrolled class
        unit_rows = []
        for enr in t['enrolled_classes']:
            cu = (db.table('class_units')
                  .select('units(id, name, code)')
                  .eq('class_id', enr['id'])
                  .execute().data or [])
            unit_rows.extend([r['units'] for r in cu if r.get('units')])
        # Deduplicate by unit id
        seen = set()
        t['units'] = []
        for u in unit_rows:
            if u['id'] not in seen:
                seen.add(u['id'])
                t['units'].append(u)

    return render_template('dept_admin/trainees.html',
                           trainees=data, departments=deps, classes=classes)


@dept_admin_bp.route('/trainees/add', methods=['POST'])
@login_required
@dept_admin_required
def add_trainee():
    db       = get_db()
    dep_id   = request.form.get('department_id') or str(current_user.department_id)
    email    = request.form.get('email', '').strip().lower()
    name     = request.form.get('full_name', '').strip()
    adm_no   = request.form.get('admission_no', '').strip()
    class_id = request.form.get('class_id', '').strip()

    if not name or not adm_no:
        flash('Full name and admission number are required.', 'danger')
        return redirect(url_for('dept_admin.trainees'))
    if len(adm_no) != 5 or not adm_no.isdigit():
        flash('Admission number must be exactly 5 digits.', 'danger')
        return redirect(url_for('dept_admin.trainees'))

    # Email is optional — generate a placeholder if not provided
    if not email:
        email = f'{adm_no}@ttieportfolio.local'

    try:
        result = db.table('users').insert({
            'email':         email,
            'full_name':     name,
            'role':          'trainee',
            'department_id': dep_id,
            'admission_no':  adm_no,
            'password_hash': None,
            'must_change_password': False,
            'created_by':    str(current_user.id),
            'is_active':     True,
        }).execute()

        if class_id and result.data:
            try:
                db.table('enrollments').insert({
                    'trainee_id': result.data[0]['id'],
                    'class_id':   class_id
                }).execute()
            except Exception:
                pass

        log_action('CREATE_TRAINEE', 'user', None, f'{name} ({adm_no})')
        flash(f'Trainee "{name}" (ADM: {adm_no}) added. '
              f'They can now activate their account using admission number and full name.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('dept_admin.trainees'))


@dept_admin_bp.route('/trainees/csv-template')
@login_required
@dept_admin_required
def trainees_csv_template():
    """Download a blank CSV template for bulk trainee import."""
    from flask import make_response
    csv_content = (
        "admission_no,full_name,class_name,course_code,course_name,unit_code,unit_name\n"
        "12345,John Doe,ICT Class A,ICT5,Information Communication Technology,ICT101,Computer Applications\n"
        "67890,Jane Smith,,EL6,Electrical Engineering,EL601,Electrical Installation\n"
    )
    resp = make_response(csv_content)
    resp.headers['Content-Disposition'] = 'attachment; filename="trainees_courses_units_import_template.csv"'
    resp.headers['Content-Type'] = 'text/csv'
    return resp


def _normalize_import_row(row: dict) -> dict:
    aliases = {
        'admission_no': ('admission_no', 'admission no', 'admission number', 'adm_no', 'adm no', 'admission'),
        'full_name': ('full_name', 'full name', 'name', 'trainee name'),
        'class_name': ('class_name', 'class name', 'class'),
        'course_code': ('course_code', 'course code'),
        'course_name': ('course_name', 'course name'),
        'unit_code': ('unit_code', 'unit code'),
        'unit_name': ('unit_name', 'unit name'),
    }
    normalized = {}
    source = {str(k).strip().lower().replace('-', '_'): v for k, v in row.items() if k is not None}
    source.update({str(k).strip().lower(): v for k, v in row.items() if k is not None})
    for target, names in aliases.items():
        value = ''
        for name in names:
            key = name.strip().lower().replace('-', '_')
            if key in source and source[key] is not None:
                value = str(source[key]).strip()
                break
        normalized[target] = value
    return normalized


def _read_trainee_import_rows(file_storage) -> tuple[list[dict], str | None]:
    filename = (file_storage.filename or '').lower()
    if filename.endswith('.csv'):
        import csv, io
        content = file_storage.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        if not reader.fieldnames:
            return [], 'File must have a header row.'
        return [_normalize_import_row(row) for row in reader], None

    if filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return [], 'Excel import requires openpyxl. Install dependencies with: pip install -r requirements.txt'

        file_storage.stream.seek(0)
        wb = load_workbook(file_storage.stream, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return [], 'Excel file must have a header row.'
        header_names = [str(h).strip() if h is not None else '' for h in headers]
        data = []
        for values in rows:
            if not values or not any(v is not None and str(v).strip() for v in values):
                continue
            data.append(_normalize_import_row(dict(zip(header_names, values))))
        return data, None

    return [], 'Please upload a .xlsx Excel file or a .csv file.'


def _get_or_create_course(db, dep_id: str, code: str, name: str) -> tuple[str | None, bool]:
    if not code or not name or not dep_id:
        return None, False
    existing = (db.table('courses').select('id')
                .eq('department_id', dep_id)
                .eq('code', code)
                .limit(1)
                .execute().data or [])
    if existing:
        return existing[0]['id'], False
    result = db.table('courses').insert({
        'name': name,
        'code': code,
        'department_id': dep_id,
        'created_by': str(current_user.id),
    }).execute()
    return (result.data[0]['id'] if result.data else None), True


def _get_or_create_unit(db, course_id: str, code: str, name: str) -> tuple[str | None, bool]:
    if not course_id or not name:
        return None, False
    q = db.table('units').select('id').eq('course_id', course_id)
    if code:
        q = q.eq('code', code)
    else:
        q = q.eq('name', name)
    existing = q.limit(1).execute().data or []
    if existing:
        return existing[0]['id'], False
    result = db.table('units').insert({
        'name': name,
        'code': code or None,
        'course_id': course_id,
        'created_by': str(current_user.id),
    }).execute()
    return (result.data[0]['id'] if result.data else None), True


@dept_admin_bp.route('/trainees/import', methods=['POST'])
@login_required
@dept_admin_required
def import_trainees():
    """
    Bulk import trainees, courses, and units from CSV or Excel.
    Expected columns (header row required):
        admission_no, full_name
        Optional: class_name, course_code, course_name, unit_code, unit_name
    """
    db     = get_db()
    dep_id = str(current_user.department_id) if current_user.department_id else None
    f      = request.files.get('import_file') or request.files.get('csv_file')

    if not f or f.filename == '':
        flash('Please select an Excel or CSV file.', 'danger')
        return redirect(url_for('dept_admin.trainees'))

    try:
        rows, error = _read_trainee_import_rows(f)
        if error:
            flash(error, 'danger')
            return redirect(url_for('dept_admin.trainees'))
        if not rows:
            flash('Import file has no data rows.', 'danger')
            return redirect(url_for('dept_admin.trainees'))

        # Build class name → id lookup
        q_classes = db.table('classes').select('id, name')
        if dep_id:
            q_classes = q_classes.eq('department_id', dep_id)
        all_classes = q_classes.execute().data or []
        class_map   = {c['name'].strip().lower(): c['id'] for c in all_classes}

        added = skipped = errors = courses_created = units_created = class_unit_links = 0
        for row in rows:
            adm_no      = row['admission_no'].strip().zfill(5)
            full_name   = row['full_name'].strip()
            cls_name    = row['class_name'].strip()
            course_code = row['course_code'].strip().upper()
            course_name = row['course_name'].strip()
            unit_code   = row['unit_code'].strip().upper()
            unit_name   = row['unit_name'].strip()

            if not adm_no or not full_name:
                errors += 1
                continue
            if not adm_no.isdigit() or len(adm_no) != 5:
                errors += 1
                continue

            course_id = unit_id = None
            if course_code or course_name or unit_code or unit_name:
                if not course_code or not course_name:
                    errors += 1
                    continue
                course_id, created = _get_or_create_course(db, dep_id, course_code, course_name)
                courses_created += 1 if created else 0
                if unit_code or unit_name:
                    if not unit_name:
                        errors += 1
                        continue
                    unit_id, created = _get_or_create_unit(db, course_id, unit_code, unit_name)
                    units_created += 1 if created else 0

            class_id = class_map.get(cls_name.lower()) if cls_name else None
            if class_id and unit_id:
                try:
                    db.table('class_units').insert({
                        'class_id': class_id,
                        'unit_id': unit_id,
                    }).execute()
                    class_unit_links += 1
                except Exception:
                    pass

            # Skip trainee creation if already exists, but keep course/unit imports above.
            existing = (db.table('users').select('id')
                        .eq('admission_no', adm_no).execute().data)
            if existing:
                skipped += 1
                continue

            email   = f'{adm_no}@ttieportfolio.local'
            try:
                result = db.table('users').insert({
                    'email':         email,
                    'full_name':     full_name,
                    'role':          'trainee',
                    'department_id': dep_id,
                    'admission_no':  adm_no,
                    'password_hash': None,
                    'must_change_password': False,
                    'created_by':    str(current_user.id),
                    'is_active':     True,
                }).execute()

                if class_id and result.data:
                    try:
                        db.table('enrollments').insert({
                            'trainee_id': result.data[0]['id'],
                            'class_id':   class_id
                        }).execute()
                    except Exception:
                        pass
                added += 1
            except Exception:
                errors += 1

        log_action('IMPORT_TRAINEES', 'user', None,
                   f'Imported {added}, skipped {skipped}, courses {courses_created}, units {units_created}, errors {errors}')
        flash(
            f'Import complete — {added} trainee(s) added, {skipped} already existed, '
            f'{courses_created} course(s) created, {units_created} unit(s) created, '
            f'{class_unit_links} class-unit link(s), {errors} error(s).',
            'success'
        )
    except Exception as e:
        flash(f'Import failed: {e}', 'danger')

    return redirect(url_for('dept_admin.trainees'))


@dept_admin_bp.route('/trainees/delete/<trainee_id>', methods=['POST'])
@login_required
@dept_admin_required
def delete_trainee(trainee_id):
    try:
        get_db().table('users').delete().eq('id', trainee_id).execute()
        log_action('DELETE_TRAINEE', 'user', trainee_id)
        flash('Trainee deleted.', 'success')
    except Exception as e:
        flash(f'Error: {e}', 'danger')
    return redirect(url_for('dept_admin.trainees'))


# ─────────────────────────────────────────────────────────────
# API: units for a class (trainee upload form)
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/api/class-units/<class_id>')
@login_required
def api_class_units(class_id):
    db   = get_db()
    rows = (db.table('class_units').select('units(id, name)')
            .eq('class_id', class_id).execute().data or [])
    return jsonify({'units': [r['units'] for r in rows if r.get('units')]})


# ─────────────────────────────────────────────────────────────
# API: evidence for an assessment (dashboard modal)
# ─────────────────────────────────────────────────────────────
@dept_admin_bp.route('/api/assessment/<assessment_id>/evidence')
@login_required
@dept_admin_required
def api_assessment_evidence(assessment_id):
    from app.utils import get_storage_public_url, STORAGE_BUCKET_EVIDENCE
    db = get_db()
    try:
        evidence = (db.table('evidence').select('*')
                    .eq('assessment_id', assessment_id)
                    .order('uploaded_at').execute().data or [])
        for ev in evidence:
            ev['url'] = get_storage_public_url(STORAGE_BUCKET_EVIDENCE, ev.get('file_path', ''))
        return jsonify({'success': True, 'evidence': evidence})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
